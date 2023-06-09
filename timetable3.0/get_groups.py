#!/usr/bin/python
# -*- coding: utf-8 -*-
import json
import os

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

is_print = True


# A   0
# B   1
# C   2
# D   3
# E   4
# F   5
# G   6
# H   7
# I   8
# J   9
# K   10
# L   11
# M   12

def get_timetable_sheet(sheet: Worksheet):
    days_of_week = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
    groups_sheet = {}
    this_group = None
    this_week_day = None
    to_pechat = False
    last_pair_number, last_pair_name, last_pair_teacher, last_pair_audience = None, None, None, None
    for row in sheet.rows:
        if 'На печать' in [cell.value for cell in row]:
            to_pechat = True
            break
    print(to_pechat)
    for row in sheet.rows:
        skip_row = False
        for cell in row:
            if cell.value or cell.value == 0:
                if cell.font.sz > 24:
                    this_group = cell.value.strip()
                    if is_print:
                        print(f"$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$\n"
                              f"$$ -> {this_group} <- $$\n"
                              f"$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
                    groups_sheet[this_group] = {}
                    skip_row = True
                    last_pair_number, last_pair_name, last_pair_teacher, last_pair_audience = None, None, None, None
                    break
                if str(cell.value).strip() in days_of_week:
                    this_week_day = cell.value.strip()
                    if is_print:
                        print(f"\n!!!{this_week_day}!!!")
                    groups_sheet[this_group][this_week_day] = {}
                    skip_row = True
                    last_pair_number, last_pair_name, last_pair_teacher, last_pair_audience = None, None, None, None
                    break
        if skip_row:
            continue
        pair_number = row[3 if to_pechat else 0].value
        pair_name = row[7 if to_pechat else 1].value
        pair_teacher = row[8 if to_pechat else 5].value
        pair_audience = row[9 if to_pechat else 8].value
        pair_is_numerator = row[7 if to_pechat else 1].fill.start_color.index in ["00000000", "0", 0]
        if pair_number is None and not pair_name and not pair_teacher and not pair_audience and pair_is_numerator:
            continue
        if pair_name == '* ДОТ - с применением дистанционных образовательных технологий':
            continue
        if pair_number is None:
            if last_pair_number is None:
                print("TOTAL ERR нет номера пары")
                continue
            else:
                pair_number = last_pair_number
        if not groups_sheet[this_group][this_week_day].get(pair_number):
            groups_sheet[this_group][this_week_day][pair_number] = {}
        if not groups_sheet[this_group][this_week_day][pair_number].get(
                'numerator' if pair_is_numerator else 'denominator'):
            if pair_name:
                groups_sheet[this_group][this_week_day][pair_number][
                    'numerator' if pair_is_numerator else 'denominator'] = {
                    'subject': '',
                    'teacher': '',
                    'audience': ''
                }
            else:
                groups_sheet[this_group][this_week_day][pair_number][
                    'numerator' if pair_is_numerator else 'denominator'] = None
                last_pair_number = pair_number
                continue
        if not pair_audience:
            pair_audience = last_pair_audience
        if not pair_name:
            pair_name = last_pair_name
        if not pair_teacher:
            pair_teacher = last_pair_teacher
        print(pair_number, 'ч' if pair_is_numerator else 'з', pair_name, pair_teacher, pair_audience)
        this_pair = groups_sheet[this_group][this_week_day][pair_number]['numerator' if pair_is_numerator else 'denominator']



        if not this_pair['subject']:
            this_pair['subject'] = pair_name
        else:
            if pair_name and this_pair['subject'].replace(' п/гр 2', '').replace(' п/гр 1', '') == pair_name.replace(' п/гр 2', '').replace(' п/гр 1', ''):
                this_pair['subject'] = pair_name.replace(' п/гр 2', '').replace(' п/гр 1', '')
            else:
                this_pair['subject'] += ' / ' + pair_name.replace(' п/гр 2', '').replace(' п/гр 1', '')

        if not this_pair['teacher']:
            this_pair['teacher'] = pair_teacher.strip().replace('\n\n', '\n').replace('\n', ' / ')
        else:
            if this_pair['teacher'] == pair_teacher:
                this_pair['teacher'] = pair_teacher.strip().replace('\n\n', '\n').replace('\n', ' / ')
            else:
                this_pair['teacher'] += ' / ' + pair_teacher.strip().replace('\n\n', '\n').replace('\n', ' / ')

        if not this_pair['audience']:
            this_pair['audience'] = pair_audience.strip().replace('\n\n', '\n').replace('\n', ' / ')
        elif pair_audience:
            if this_pair['audience'] == pair_audience.strip().replace('\n\n', '\n').replace('\n', ' / '):
                this_pair['audience'] = pair_audience.strip().replace('\n\n', '\n').replace('\n', ' / ')
            else:
                this_pair['audience'] += ' / ' + pair_audience.strip().replace('\n\n', '\n').replace('\n', ' / ')

        last_pair_number, last_pair_name, last_pair_teacher, last_pair_audience = pair_number, pair_name, pair_teacher, pair_audience
        # if False:
        #     # print([cell.value for cell in row])
        #     for cell in row:
        #         if cell.value or cell.value == 0:
        #             if cell.font.sz > 24:
        #                 this_group = cell.value.strip()
        #                 if is_print:
        #                     print(f"\n$$$$$$$$$$$$$$$$$$$$$$$$$$$$$\n"
        #                           f"$$$$ -> {this_group} <- $$$$\n"
        #                           f"$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
        #                 groups_sheet[this_group] = {}
        #             elif str(cell.value).strip() in days_of_week:
        #                 this_week_day = cell.value.strip()
        #                 if is_print:
        #                     print(f"\n!!!{this_week_day}!!!")
        #                 groups_sheet[this_group][this_week_day] = {}
        #             elif str(cell.value).isdigit() and 'A' in cell.coordinate:
        #                 number_lesson = int(cell.value)
        #                 row_number = int(cell.coordinate[1:])
        #                 subject = sheet[f'B{row_number}'].value
        #                 if subject:
        #                     subject = subject.replace("\n", " ").strip()
        #                 subject_zn = sheet[f'B{row_number + 1}'].value
        #                 if subject_zn:
        #                     subject_zn = subject_zn.replace("\n", " ").strip()
        #                 teacher = sheet[f'F{row_number}'].value
        #                 if teacher:
        #                     teacher = teacher.replace("\n", " ").strip()
        #                 teacher_zn = sheet[f'F{row_number + 1}'].value
        #                 if teacher_zn:
        #                     teacher_zn = teacher_zn.replace("\n", " ").strip()
        #                 audience = sheet[f'I{row_number}'].value
        #                 if audience:
        #                     audience = audience.replace("\n", " ").strip()
        #                 audience_zn = sheet[f'I{row_number + 1}'].value
        #                 if audience_zn:
        #                     audience_zn = audience_zn.replace("\n", " ").strip()
        #                 if teacher_zn is None and subject_zn:
        #                     teacher_zn = teacher
        #                 if audience_zn is None and subject_zn:
        #                     audience_zn = audience
        #                 if not subject and not subject_zn:
        #                     continue
        #                 if is_print:
        #                     print(f"[{number_lesson}] {subject} {teacher} {audience}\n"
        #                           f"{subject_zn} {teacher_zn} {audience_zn} color: ",
        #                           sheet[f'B{row_number + 1}'].fill.start_color.index)
        #                 #
        #                 if sheet[f'B{row_number + 1}'].fill.start_color.index != "00000000" \
        #                         and subject_zn is None:  # Только числитель
        #                     if is_print:
        #                         print("Только числитель")
        #                     groups_sheet[this_group][this_week_day][number_lesson] = {
        #                         "numerator": {
        #                             "subject": subject,
        #                             "teacher": teacher,
        #                             "audience": audience,
        #                         }
        #                     }
        #                 elif subject is None and subject_zn:  # Только знаменатель
        #                     if is_print:
        #                         print("Только знаменатель")
        #                     groups_sheet[this_group][this_week_day][number_lesson] = {
        #                         "denominator": {
        #                             "subject": subject_zn,
        #                             "teacher": teacher_zn,
        #                             "audience": audience_zn,
        #                         },
        #                     }
        #                 elif subject and not subject_zn:  # Числитель и Знаменатель одинаковы
        #                     if is_print:
        #                         print("Числитель и Знаменатель одинаковы")
        #                     groups_sheet[this_group][this_week_day][number_lesson] = {
        #                         "numerator": {
        #                             "subject": subject,
        #                             "teacher": teacher,
        #                             "audience": audience,
        #                         },
        #                         "denominator": {
        #                             "subject": subject,
        #                             "teacher": teacher,
        #                             "audience": audience,
        #                         }
        #                     }
        #                 elif subject and subject_zn:  # Числитель и Знаменатель разные
        #                     if is_print:
        #                         print("Числитель и Знаменатель разные")
        #                     groups_sheet[this_group][this_week_day][number_lesson] = {
        #                         "numerator": {
        #                             "subject": subject,
        #                             "teacher": teacher,
        #                             "audience": audience,
        #                         },
        #                         "denominator": {
        #                             "subject": subject_zn,
        #                             "teacher": teacher_zn,
        #                             "audience": audience_zn,
        #                         }
        #                     }
        #                 else:
        #                     print("TOTAL ERR", cell)
        #                     print(subject, teacher, audience)
        #                     print(subject_zn, teacher_zn, audience_zn)
    for current_group in groups_sheet:
        for current_day_of_week in groups_sheet[current_group]:
            for current_pair_number in groups_sheet[current_group][current_day_of_week]:
                this_pair = groups_sheet[current_group][current_day_of_week][current_pair_number]
                if len(this_pair.keys()) == 1:
                    groups_sheet[current_group][current_day_of_week][current_pair_number]['denominator'] = this_pair[
                        'numerator']
                else:
                    if not groups_sheet[current_group][current_day_of_week][current_pair_number]['denominator']:
                        groups_sheet[current_group][current_day_of_week][current_pair_number].pop('denominator')
                    if not groups_sheet[current_group][current_day_of_week][current_pair_number]['numerator']:
                        groups_sheet[current_group][current_day_of_week][current_pair_number].pop('numerator')
    return groups_sheet


def get_timetable_departament(departament):
    groups_departament = {}
    workbook = load_workbook("groups/" + departament)
    for sheet_name in workbook.sheetnames:
        print(sheet_name)
        groups_departament[sheet_name.strip()] = get_timetable_sheet(workbook[sheet_name])
    return groups_departament


def get_all_groups():
    groups = {}
    for departament_name in os.listdir(path="groups"):
        if "$" in departament_name or '.xlsx' not in departament_name:
            continue
        print("Отделение:", departament_name)
        this_departament = get_timetable_departament(departament_name)
        if len(this_departament.keys()) < 10:
            groups[departament_name.replace(".xlsx", "")] = this_departament
        else:
            first_this_departament = {}
            for to_add_dep in list(this_departament.keys())[int(len(this_departament) / 2):]:
                first_this_departament[to_add_dep] = this_departament[to_add_dep]
                del this_departament[to_add_dep]
            groups[departament_name.replace(".xlsx", "") + " 1 Часть"] = this_departament
            groups[departament_name.replace(".xlsx", "") + " 2 Часть"] = first_this_departament
    with open('groups.txt', 'w', encoding='utf-8') as f:
        f.write(json.dumps(groups, ensure_ascii=False))
    # print(json.dumps(groups, ensure_ascii=False))


# workbook = load_workbook("groups/" + "ОАР.xlsx")
# print(get_timetable_sheet(workbook[workbook.sheetnames[4]]))
get_all_groups()
